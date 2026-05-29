import pandas as pd
import io
import sys
import os
from openpyxl import Workbook

# Añadir el path del backend para importar el extractor
sys.path.append(os.path.join(os.getcwd(), 'backend_v2'))

from app.services.novedades_nomina.control_descuentos_extractor import extraer_control_descuentos

def test_priority():
    print("Iniciando prueba de PRIORIDAD de columnas para CONTROL DE DESCUENTOS...")
    
    # 1. Crear un Workbook con columnas conflictivas
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    
    # Escribir encabezados en FILA 6 (1-indexed)
    ws.cell(row=6, column=1, value="ID")          # Columna que antes causaba conflicto (ej: index 1, 2, 3)
    ws.cell(row=6, column=2, value="NOMBRE")      # Columna que contiene Cedula - Nombre
    ws.cell(row=6, column=3, value="VALOR CUOTA")
    
    # Escribir datos en FILA 7+
    # ID: 1, 2, 3...
    # NOMBRE: "101010 - JUAN PEREZ"
    data = [
        (1, '101010 - JUAN PEREZ', 150.00),
        (2, '202020 - MARIA LOPEZ', 250.00),
        (3, '101010 - JUAN PEREZ', 50.00)  # Repetido para probar suma
    ]
    
    for i, (idx, name, val) in enumerate(data):
        ws.cell(row=7+i, column=1, value=idx)
        ws.cell(row=7+i, column=2, value=name)
        ws.cell(row=7+i, column=3, value=val)
    
    output = io.BytesIO()
    wb.save(output)
    excel_binario = output.getvalue()
    
    # 2. Ejecutar extractor
    rows, summary, warnings = extraer_control_descuentos([excel_binario])
    
    print(f"Summary: {summary}")
    print(f"Warnings: {warnings}")
    
    # 3. Verificar resultados
    # Se espera que ignore la columna ID (1,2,3) y use NOMBRE para la cedula
    # esperado: 101010: 200.0, 202020: 250.0
    expected = {
        '101010': 200.00,
        '202020': 250.00
    }
    
    success = True
    if len(rows) != 2:
        print(f"ERROR: Se esperaban 2 registros únicos, se obtuvieron {len(rows)}")
        success = False
    
    for r in rows:
        ced = r['cedula']
        val = r['valor']
        if ced in expected:
            if abs(val - expected[ced]) > 0.001:
                print(f"ERROR: Valor incorrecto para {ced}. Esperado: {expected[ced]}, Obtenido: {val}")
                success = False
            else:
                print(f"OK: Cédula {ced} procesada correctamente con valor {val}")
        else:
            print(f"ERROR: Cédula inesperada {ced}. Probablemente tomó el ID de la fila en lugar de la cédula.")
            success = False
            
    if success:
        print("\nPRUEBA EXITOSA: La prioridad de columnas funciona correctamente.")
    else:
        print("\nPRUEBA FALLIDA.")

if __name__ == "__main__":
    test_priority()

import pandas as pd
import io
import sys
import os
from openpyxl import Workbook

# Añadir el path del backend para importar el extractor
sys.path.append(os.path.join(os.getcwd(), 'backend_v2'))

from app.services.novedades_nomina.control_descuentos_extractor import extraer_control_descuentos

def test_extractor():
    print("Iniciando prueba de extractor CONTROL DE DESCUENTOS...")
    
    # 1. Crear un Workbook manual con openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    
    # Escribir algo en las primeras filas para que no estén "vacías"
    for r in range(1, 6):
        ws.cell(row=r, column=1, value=f"BASURA {r}")
        ws.cell(row=r, column=2, value=f"BASURA {r}")

    # Escribir encabezados en FILA 6 (1-indexed)
    ws.cell(row=6, column=1, value="CEDULA")
    ws.cell(row=6, column=2, value="VALOR CUOTA")
    
    # Escribir datos en FILA 7+
    data = [
        ('123456', 100.50),
        ('789012', 200.00),
        ('123456', 50.25),
        ('111222', 300.75),
        ('789012', 100.00)
    ]
    
    for i, (ced, val) in enumerate(data):
        ws.cell(row=7+i, column=1, value=ced)
        ws.cell(row=7+i, column=2, value=val)
    
    output = io.BytesIO()
    wb.save(output)
    excel_binario = output.getvalue()
    
    # DEBUG PROFUNDO
    print("\n--- DEBUG PROFUNDO ---")
    try:
        df_raw = pd.read_excel(io.BytesIO(excel_binario), sheet_name='Hoja1', header=None)
        print("Raw Excel - Fila 6 (0-indexed 5):")
        print(df_raw.iloc[5].tolist())
        
        df_skip = pd.read_excel(io.BytesIO(excel_binario), sheet_name='Hoja1', skiprows=5)
        print("\nSkiprows=5 columns:")
        print(df_skip.columns.tolist())
        print("\nSkiprows=5 content:")
        print(df_skip.head())
        
    except Exception as e:
        print(f"DEBUG ERROR: {e}")
    print("----------------------\n")

    # 2. Ejecutar extractor
    rows, summary, warnings = extraer_control_descuentos([excel_binario])
    
    print(f"Summary: {summary}")
    print(f"Warnings: {warnings}")
    
    # 3. Verificar resultados
    expected = {
        '123456': 150.75,
        '789012': 300.00,
        '111222': 300.75
    }
    
    success = True
    if len(rows) != 3:
        print(f"ERROR: Se esperaban 3 filas agrupadas, se obtuvieron {len(rows)}")
        success = False
        
    for r in rows:
        ced = r['cedula']
        val = r['valor']
        if ced in expected:
            if abs(val - expected[ced]) > 0.001:
                print(f"ERROR: Valor incorrecto para {ced}. Esperado: {expected[ced]}, Obtenido: {val}")
                success = False
        else:
            print(f"ERROR: Cédula inesperada {ced}")
            success = False
            
    if success:
        print("PRUEBA EXITOSA: El extractor funciona correctamente.")
    else:
        print("PRUEBA FALLIDA: Revisar errores arriba.")

if __name__ == "__main__":
    test_extractor()

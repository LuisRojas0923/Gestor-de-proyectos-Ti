"""Validación estructural y de recursos para archivos Excel de Seguros HDI."""

import io
import zipfile
from pathlib import Path, PurePosixPath

import openpyxl
import xlrd


MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024
MAX_EXCEL_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
MAX_EXCEL_MEMBER_BYTES = 20 * 1024 * 1024
MAX_EXCEL_ENTRIES = 1000
MAX_COMPRESSION_RATIO = 100
MAX_EXCEL_SHEETS = 10
MAX_EXCEL_ROWS = 5000
MAX_EXCEL_COLS = 50

MIME_TYPES = {
    ".xlsx": {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/octet-stream",
        "application/zip",
        "",
    },
    ".xls": {
        "application/vnd.ms-excel",
        "application/octet-stream",
        "",
    },
}


class ArchivoHdiInvalido(ValueError):
    """El archivo incumple el contrato de seguridad de Seguros HDI."""

    def __init__(self, detail: str, status_code: int = 415):
        super().__init__(detail)
        self.detail = detail
        self.status_code = status_code


def _validar_limites_dimensiones(dimensiones: list[tuple[int, int]]) -> None:
    if len(dimensiones) > MAX_EXCEL_SHEETS:
        raise ArchivoHdiInvalido(
            f"El libro supera el límite de {MAX_EXCEL_SHEETS} hojas.", 413
        )
    for filas, columnas in dimensiones:
        if filas > MAX_EXCEL_ROWS:
            raise ArchivoHdiInvalido(
                f"Una hoja supera el límite de {MAX_EXCEL_ROWS} filas.", 413
            )
        if columnas > MAX_EXCEL_COLS:
            raise ArchivoHdiInvalido(
                f"Una hoja supera el límite de {MAX_EXCEL_COLS} columnas.", 413
            )


def _validar_xlsx(contenido: bytes) -> None:
    if not contenido.startswith(b"PK\x03\x04"):
        raise ArchivoHdiInvalido("La extensión .xlsx no coincide con el contenido.")

    try:
        with zipfile.ZipFile(io.BytesIO(contenido)) as archive:
            entries = archive.infolist()
            if len(entries) > MAX_EXCEL_ENTRIES:
                raise ArchivoHdiInvalido("El XLSX contiene demasiadas entradas.", 413)

            names = {entry.filename for entry in entries}
            required = {
                "[Content_Types].xml",
                "_rels/.rels",
                "xl/workbook.xml",
                "xl/_rels/workbook.xml.rels",
            }
            worksheets = {
                name for name in names
                if name.startswith("xl/worksheets/") and name.endswith(".xml")
            }
            if not required.issubset(names) or not worksheets:
                raise ArchivoHdiInvalido("El contenido no corresponde a un libro XLSX.")

            total_uncompressed = 0
            for entry in entries:
                path = PurePosixPath(entry.filename)
                if path.is_absolute() or ".." in path.parts or "\\" in entry.filename:
                    raise ArchivoHdiInvalido("El XLSX contiene rutas internas no válidas.")
                if entry.flag_bits & 0x1:
                    raise ArchivoHdiInvalido("No se permiten libros cifrados.")
                if entry.file_size > MAX_EXCEL_MEMBER_BYTES:
                    raise ArchivoHdiInvalido("Una entrada del XLSX es demasiado grande.", 413)
                total_uncompressed += entry.file_size
                if total_uncompressed > MAX_EXCEL_UNCOMPRESSED_BYTES:
                    raise ArchivoHdiInvalido("El XLSX descomprimido es demasiado grande.", 413)
                ratio = entry.file_size / max(entry.compress_size, 1)
                if ratio > MAX_COMPRESSION_RATIO:
                    raise ArchivoHdiInvalido("El XLSX tiene una compresión no segura.", 413)

            # Leer cada entrada obliga a zipfile a comprobar CRC sin extraer a disco.
            for entry in entries:
                with archive.open(entry) as source:
                    while source.read(1024 * 1024):
                        pass
    except ArchivoHdiInvalido:
        raise
    except (OSError, RuntimeError, zipfile.BadZipFile, zipfile.LargeZipFile) as exc:
        raise ArchivoHdiInvalido("El archivo XLSX está corrupto.") from exc

    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(contenido), read_only=True, data_only=True
        )
        try:
            _validar_limites_dimensiones([
                (sheet.max_row or 0, sheet.max_column or 0)
                for sheet in workbook.worksheets
            ])
        finally:
            workbook.close()
    except ArchivoHdiInvalido:
        raise
    except Exception as exc:
        raise ArchivoHdiInvalido("El archivo XLSX no puede abrirse.") from exc


def _validar_xls(contenido: bytes) -> None:
    ole_signature = bytes.fromhex("D0CF11E0A1B11AE1")
    if not contenido.startswith(ole_signature):
        raise ArchivoHdiInvalido("La extensión .xls no coincide con el contenido.")
    try:
        workbook = xlrd.open_workbook(file_contents=contenido, on_demand=True)
        try:
            _validar_limites_dimensiones([
                (workbook.sheet_by_index(index).nrows, workbook.sheet_by_index(index).ncols)
                for index in range(workbook.nsheets)
            ])
        finally:
            workbook.release_resources()
    except ArchivoHdiInvalido:
        raise
    except Exception as exc:
        raise ArchivoHdiInvalido("El archivo XLS está corrupto o no es un libro BIFF válido.") from exc


def validar_excel_hdi(
    contenido: bytes,
    nombre_archivo: str,
    tipo_mime: str | None = None,
) -> str:
    """Valida el archivo completo y devuelve su extensión real sin punto."""
    nombre = Path((nombre_archivo or "").replace("\\", "/")).name
    extension = Path(nombre).suffix.lower()
    if extension not in MIME_TYPES:
        raise ArchivoHdiInvalido("Solo se permiten archivos Excel .xls o .xlsx.")
    if not contenido:
        raise ArchivoHdiInvalido("El archivo está vacío.", 400)
    if len(contenido) > MAX_FILE_SIZE_BYTES:
        raise ArchivoHdiInvalido("El archivo supera el límite de 10 MB.", 413)
    if tipo_mime is not None and tipo_mime.lower() not in MIME_TYPES[extension]:
        raise ArchivoHdiInvalido("El tipo MIME no coincide con la extensión del archivo.")

    if extension == ".xlsx":
        _validar_xlsx(contenido)
    else:
        _validar_xls(contenido)
    return extension.removeprefix(".")

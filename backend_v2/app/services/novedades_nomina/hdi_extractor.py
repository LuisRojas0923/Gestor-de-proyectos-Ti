"""
Extractor especializado para planillas Excel de SEGUROS HDI.
"""
import io
import re
import collections
import logging
from typing import List, Dict, Any, Tuple
import pandas as pd

logger = logging.getLogger(__name__)

def _limpiar_numero(val: Any) -> float:
    """Limpia caracteres de moneda y separadores."""
    if pd.isna(val):
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)

    s = str(val).replace("$", "").replace(" ", "").replace(",", "")
    try:
        return float(s)
    except:
        return 0.0

def _aplicar_reemplazos(cedula: str, nombre: str) -> Tuple[str, str]:
    """Aplica las reglas específicas de reemplazo de HDI."""
    # IDs
    if cedula == "1116235786": cedula = "66903320"
    if cedula == "31282865": cedula = "31231202"

    # Nombres
    n_up = nombre.upper()
    if "HECTOR PAUL CRUZ" in n_up: nombre = "MARIBEL TORRES AGUDELO"
    if "ESPERANZA AGUADO CORTES" in n_up: nombre = "GLORIA AGUDELO DE TORRES"

    return cedula, nombre

def _formatear_nombre(n: str) -> str:
    # Eliminar dobles espacios y capitalizar
    return " ".join(n.split()).title()

def extraer_hdi(
    archivos_binarios: List[bytes]
) -> Tuple[List[Dict[str, Any]], Dict[str, Any], List[str]]:
    """Procesa planillas Excel de HDI (en lugar del PDF antiguo).

    Límites OOXML aplicados antes del parsing completo:
    - MAX_SHEETS_POR_ARCHIVO: máximo de hojas por archivo
    - MAX_FILAS_POR_HOJA: máximo de filas por hoja para prevenir agotamiento de memoria
    """
    MAX_SHEETS_POR_ARCHIVO = 5
    MAX_FILAS_POR_HOJA = 50_000

    all_raw_rows: List[Dict[str, Any]] = []
    warnings: List[str] = []

    for contenido in archivos_binarios:
        try:
            # Validar límites OOXML antes del parsing completo (modo read_only evita cargar DOM entero)
            import openpyxl
            wb_check = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
            n_sheets = len(wb_check.sheetnames)
            if n_sheets > MAX_SHEETS_POR_ARCHIVO:
                warnings.append(f"Archivo con {n_sheets} hojas excede el límite de {MAX_SHEETS_POR_ARCHIVO}. Se omite.")
                wb_check.close()
                continue

            # Verificar filas por hoja
            excede_filas = False
            for sheet_name in wb_check.sheetnames:
                ws = wb_check[sheet_name]
                if ws.max_row is not None and ws.max_row > MAX_FILAS_POR_HOJA:
                    warnings.append(
                        f"Hoja '{sheet_name}' tiene {ws.max_row} filas, "
                        f"excede el límite de {MAX_FILAS_POR_HOJA}. Se omite el archivo."
                    )
                    excede_filas = True
                    break
            wb_check.close()

            if excede_filas:
                continue

            # Leer el Excel omitiendo la primera fila de título "RELACION DE ASEGURADOS"
            df = pd.read_excel(io.BytesIO(contenido), skiprows=1)

            # Eliminar la primera columna si viene vacía (Unnamed: 0)
            if df.columns[0].startswith("Unnamed:"):
                df = df.drop(columns=[df.columns[0]])

            # Identificar la columna de PRIMA COBRO (que tiene celdas combinadas) y hacer forward fill
            # Aunque no se use en la lógica matemática interna (porque se recalcula), es bueno normalizarla
            col_cobro = None
            for col in df.columns:
                if "COBRO" in str(col).upper():
                    col_cobro = col
                    break

            if col_cobro:
                df[col_cobro] = df[col_cobro].ffill()

            # Procesar las filas
            for _, row in df.iterrows():
                # Encontrar dinámicamente el nombre de la columna TIPO, IDENTIFICACION, NOMBRES, PRIMA ANUAL, CERT
                col_cert = col_tipo = col_id = col_nombre = col_prima = None
                for col in df.columns:
                    cu = str(col).upper()
                    if "CERT" in cu: col_cert = col
                    elif "TIPO" in cu: col_tipo = col
                    elif "IDENTIFICACION" in cu: col_id = col
                    elif "NOMBRE" in cu: col_nombre = col
                    elif "PRIMA ANUAL" in cu: col_prima = col

                # Fallback por índice si los nombres de columna fallan
                if not col_cert: col_cert = df.columns[0]
                if not col_tipo: col_tipo = df.columns[1] if len(df.columns) > 1 else None
                if not col_id: col_id = df.columns[2] if len(df.columns) > 2 else None
                if not col_nombre: col_nombre = df.columns[3] if len(df.columns) > 3 else None
                if not col_prima: col_prima = df.columns[-3] if len(df.columns) >= 3 else None

                # Extraer datos de la fila
                tipo = str(row.get(col_tipo, "")).strip().upper()
                if tipo not in ["P", "D"]:
                    continue # Saltar filas inválidas o totales

                raw_id = str(row.get(col_id, "")).strip().split(".")[0]
                id_val = re.sub(r"[^0-9]", "", raw_id)
                if not id_val or len(id_val) < 5:
                    continue

                nombre = str(row.get(col_nombre, "")).strip()
                prima_anual = _limpiar_numero(row.get(col_prima, 0))

                cert_raw = str(row.get(col_cert, "")).strip().split(".")[0]
                cert_val = re.sub(r"[^0-9]", "", cert_raw)

                id_val, nombre = _aplicar_reemplazos(id_val, nombre)
                nombre = _formatear_nombre(nombre)

                if prima_anual > 0:
                    all_raw_rows.append({
                        "cert": cert_val,
                        "tipo": tipo,
                        "cedula": id_val,
                        "nombre_asociado": nombre,
                        "empresa": "REFRIDCOL",
                        "prima_anual": prima_anual,
                        "concepto": "SEGURO DE VIDA"
                    })

        except Exception as e:
            logger.error(f"Error procesando HDI Excel: {e}")
            warnings.append(str(e))

    # Agrupar por CERT
    grupos_cert = collections.defaultdict(list)
    for idx, r in enumerate(all_raw_rows):
        cert_val = r.get("cert")
        if not cert_val:
            cert_val = f"dummy_{idx}"
        grupos_cert[cert_val].append(r)

    consolidated_rows: List[Dict[str, Any]] = []
    for cert_val, members in grupos_cert.items():
        # Encontrar el titular (tipo == "P")
        titular = None
        for m in members:
            if m["tipo"] == "P":
                titular = m
                break

        # Si no hay titular, usamos el primer miembro del grupo y generamos una advertencia
        if not titular:
            titular = members[0]
            warnings.append(
                f"No se detectó un Titular (TIPO = 'P') para el grupo CERT '{cert_val}'. "
                f"Se asumirá como titular a '{titular['nombre_asociado']}' y no se aplicará subsidio de empresa."
            )
            total_empleado = sum(m["prima_anual"] / 12 for m in members)
            valor_colaborador = round(total_empleado, 2)
            valor_rdc = 0.0
            valor_total = valor_colaborador
        else:
            # Hay titular:
            # 1. Prima titular: 76% colaborador, 24% empresa
            prima_titular = titular["prima_anual"] / 12
            valor_rdc = round(prima_titular * 0.24, 2)
            valor_col_titular = round(prima_titular * 0.76, 2)

            # 2. Prima dependientes: 100% colaborador, 0% empresa
            valor_col_dependents = sum(m["prima_anual"] / 12 for m in members if m is not titular)

            valor_colaborador = round(valor_col_titular + valor_col_dependents, 2)
            valor_total = round(valor_rdc + valor_colaborador, 2)

        # Crear la fila consolidada asociada al Titular
        obs = f"Grupo CERT {cert_val}"
        if len(members) > 1:
            nombres_dep = ", ".join(m["nombre_asociado"] for m in members if m is not titular)
            obs += f" | Incluye dependientes: {nombres_dep}"

        consolidated_rows.append({
            "cedula": titular["cedula"],
            "nombre_asociado": titular["nombre_asociado"],
            "empresa": titular.get("empresa", "REFRIDCOL"),
            "concepto": titular["concepto"],
            "valor": valor_total,
            "valor_rdc": valor_rdc,
            "valor_colaborador": valor_colaborador,
            "observaciones": obs
        })

    # CONSOLIDACIÓN FINAL: Agrupar por cédula (si un Titular tiene más de un CERT)
    consolidated_dict: Dict[str, Dict[str, Any]] = {}
    for r in consolidated_rows:
        ced = r["cedula"]
        if ced not in consolidated_dict:
            consolidated_dict[ced] = r
        else:
            consolidated_dict[ced]["valor"] += r["valor"]
            consolidated_dict[ced]["valor_rdc"] += r["valor_rdc"]
            consolidated_dict[ced]["valor_colaborador"] += r["valor_colaborador"]
            consolidated_dict[ced]["observaciones"] += f" | {r['observaciones']}"

            # Redondeos para evitar muchos decimales
            consolidated_dict[ced]["valor"] = round(consolidated_dict[ced]["valor"], 2)
            consolidated_dict[ced]["valor_rdc"] = round(consolidated_dict[ced]["valor_rdc"], 2)
            consolidated_dict[ced]["valor_colaborador"] = round(consolidated_dict[ced]["valor_colaborador"], 2)

    final_rows = list(consolidated_dict.values())

    # Prevenir que haya filas sin valor_colaborador (debería venir del consolidation final, pero por si acaso)
    for r in final_rows:
        if "valor_colaborador" not in r:
            r["valor_colaborador"] = r["valor"]

    total_descuentos = sum(r["valor"] for r in final_rows)
    summary = {
        "total_registros": len(final_rows),
        "total_descuentos": total_descuentos
    }

    return final_rows, summary, warnings

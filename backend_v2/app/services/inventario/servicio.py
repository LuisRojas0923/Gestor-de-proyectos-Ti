import openpyxl
from io import BytesIO
from typing import Dict, Any
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text, and_
from sqlmodel import select
from ...models.inventario.conteo import ConteoInventario, TransitoInventario, ConteoHistorico

class ServicioInventario:
    @staticmethod
    async def crear_snapshot(db: AsyncSession):
        """Copia todos los registros actuales a la tabla de historial antes de limpiar."""
        stmt = select(ConteoInventario)
        result = await db.execute(stmt)
        items = result.scalars().all()
        
        for item in items:
            hist = ConteoHistorico(
                original_id=item.id,
                b_siigo=item.b_siigo,
                bodega=item.bodega,
                bloque=item.bloque,
                estante=item.estante,
                nivel=item.nivel,
                codigo=item.codigo,
                descripcion=item.descripcion,
                unidad=item.unidad,
                cantidad_sistema=item.cantidad_sistema,
                cant_c1=item.cant_c1,
                cant_c2=item.cant_c2,
                cant_c3=item.cant_c3,
                cant_c4=item.cant_c4,
                conteo=item.conteo,
                estado=item.estado,
                invporlegalizar=item.invporlegalizar,
                cantidad_final=item.cantidad_final,
                diferencia_total=item.diferencia_total
            )
            db.add(hist)
        
        await db.flush() # Asegurar que se envíen al historico

    @staticmethod
    async def importar_conteo_excel(
        file_content: bytes, 
        nombre_conteo: str,
        db: AsyncSession,
        limpiar_previo: bool = False
    ) -> Dict[str, Any]:
        """
        Procesa un archivo Excel y carga los registros en inventario_conteo.
        Se asume que el Excel tiene las columnas en el orden de la imagen:
        B. Siigo | Bodega | Bloque | Estante | Nivel | Codigo | Descripcion | Und. | Cant. | Observaciones
        """
        wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        sheet = wb.active
        
        registros_creados = 0
        errores = []

        # 0. Si se solicita limpiar, primero respaldamos
        if limpiar_previo:
            await ServicioInventario.crear_snapshot(db)
            await db.execute(text("TRUNCATE TABLE conteoinventario RESTART IDENTITY CASCADE"))

        # Iterar desde la segunda fila (asumiendo cabecera)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            
            try:
                # Datos del Excel con limpieza de espacios
                bodega_val = str(row[1]).strip() if row[1] is not None else ""
                bloque_val = str(row[2]).strip() if row[2] is not None else ""
                estante_val = str(row[3]).strip() if row[3] is not None else ""
                nivel_val = str(row[4]).strip() if row[4] is not None else ""
                codigo_val = str(row[5]).strip() if row[5] is not None else ""
                descripcion_val = str(row[6]).strip() if row[6] is not None else ""
                cant_sist_val = float(row[8]) if row[8] is not None else 0.0
                legalizar_val = float(row[9]) if len(row) > 9 and row[9] is not None else 0.0

                if not limpiar_previo:
                    # Lógica UPSERT: Buscar si ya existe el SKU en esta ubicación
                    stmt = select(ConteoInventario).where(and_(
                        ConteoInventario.codigo == codigo_val,
                        ConteoInventario.bodega == bodega_val,
                        ConteoInventario.bloque == bloque_val,
                        ConteoInventario.estante == estante_val,
                        ConteoInventario.nivel == nivel_val
                    ))
                    res = await db.execute(stmt)
                    item_existente = res.scalar_one_or_none()
                    
                    if item_existente:
                        # Actualizar solo saldos teóricos y metadata
                        item_existente.cantidad_sistema = cant_sist_val
                        item_existente.invporlegalizar = legalizar_val
                        cant_f = cant_sist_val + legalizar_val
                        item_existente.cantidad_final = cant_f
                        # Si no ha sido contado, la diferencia total es el negativo del teórico
                        if item_existente.cant_c1 is None and item_existente.cant_c2 is None:
                            item_existente.diferencia_total = -cant_f
                            
                        item_existente.conteo = nombre_conteo
                        item_existente.descripcion = descripcion_val
                        item_existente.unidad = str(row[7]) if row[7] is not None else ""
                        
                        # Si estaba CONCILIADO, lo regresamos a PENDIENTE si el teórico cambió
                        if item_existente.estado == "CONCILIADO":
                            item_existente.estado = "PENDIENTE"
                        registros_creados += 1
                        continue

                # Crear nuevo si no existe o si limpiamos previo
                conteo = ConteoInventario(
                    b_siigo=int(row[0]) if row[0] is not None and str(row[0]).isdigit() else None,
                    bodega=bodega_val,
                    bloque=bloque_val,
                    estante=estante_val,
                    nivel=nivel_val,
                    codigo=codigo_val,
                    descripcion=descripcion_val,
                    unidad=str(row[7]) if row[7] is not None else "",
                    cantidad_sistema=cant_sist_val,
                    invporlegalizar=legalizar_val,
                    cantidad_final=cant_sist_val + legalizar_val,
                    diferencia_total=-(cant_sist_val + legalizar_val),
                    conteo=nombre_conteo
                )
                
                db.add(conteo)
                registros_creados += 1
            except Exception as e:
                errores.append(f"Fila {row_idx}: {str(e)}")
        
        if registros_creados > 0:
            await db.commit()
            
        return {
            "exito": len(errores) == 0,
            "creados": registros_creados,
            "errores": errores
        }

    @staticmethod
    def _parse_float(val: Any) -> float:
        """Limpia y parsea números que pueden venir con comas o como strings."""
        if val is None:
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        try:
            # Reemplazar coma por punto y quitar espacios
            clean_val = str(val).replace(',', '.').replace(' ', '').strip()
            return float(clean_val)
        except (ValueError, TypeError):
            return 0.0

    @staticmethod
    async def importar_transito_excel(
        file_content: bytes, 
        db: AsyncSession
    ) -> Dict[str, Any]:
        """
        Procesa un archivo Excel de tránsito (Modelo B).
        Estructura esperada por usuario: Codigo / SKU (0) | Documento (1) | Cantidad (2)
        """
        wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        sheet = wb.active
        
        registros_creados = 0
        errores = []
        
        # 1. Limpiar tabla previa de tránsito (Se asume carga fresca)
        await db.execute(text("TRUNCATE TABLE transitoinventario"))

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            
            try:
                # Mapeo según archivo usuario: 0: SKU/Codigo, 1: Documento, 2: Cantidad
                sku_val = str(row[0]).strip() if row[0] is not None else ""
                doc_val = str(row[1]).strip() if len(row) > 1 and row[1] is not None else "TRANSITO_S_D"
                cant_val = ServicioInventario._parse_float(row[2]) if len(row) > 2 else 0.0

                transito = TransitoInventario(
                    sku=sku_val,
                    documento=doc_val,
                    cantidad=cant_val
                )
                db.add(transito)
                registros_creados += 1
                
            except Exception as e:
                errores.append(f"Fila {row_idx}: {str(e)}")
        
        if registros_creados > 0:
            await db.commit()
            # 2. Sincronizar automáticamente con la tabla maestra
            await ServicioInventario.sincronizar_transito_con_maestra(db)
            
        return {
            "exito": len(errores) == 0,
            "creados": registros_creados,
            "errores": errores
        }

    @staticmethod
    async def sincronizar_transito_con_maestra(db: AsyncSession):
        """
        Calcula la suma de tránsito por SKU y actualiza la tabla conteoinventario globalmente.
        """
        # 1. Agrupar tránsito por SKU en la DB
        q_sum = text("""
            SELECT sku, SUM(cantidad) as total_transito
            FROM transitoinventario
            GROUP BY sku
        """)
        res_sum = await db.execute(q_sum)
        transito_dict = {row[0]: float(row[1]) for row in res_sum.fetchall()}

        # 2. Resetear todos los 'invporlegalizar' a 0 antes de aplicar los nuevos
        await db.execute(text("UPDATE conteoinventario SET invporlegalizar = 0"))

        # 3. Actualizar registros existentes en maestra
        # Nota: Se hace por SKU, puede haber múltiples ubicaciones por SKU
        for sku, total in transito_dict.items():
            # Actualizar todos los registros que coincidan con el SKU
            stmt = text("""
                UPDATE conteoinventario 
                SET invporlegalizar = :total,
                    cantidad_final = cantidad_sistema + :total,
                    diferencia_total = -(cantidad_sistema + :total)
                WHERE codigo = :sku
            """)
            await db.execute(stmt, {"total": total, "sku": sku})
        
        await db.commit()

    @staticmethod
    def generar_plantilla_maestra() -> BytesIO:
        """Genera un buffer con la plantilla Excel Maestra."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Maestra_Inventario"
        # Cabeceras: B. Siigo | Bodega | Bloque | Estante | Nivel | Codigo | Descripcion | Und | Cant. Sistema | Observaciones
        headers = ["B. Siigo", "Bodega", "Bloque", "Estante", "Nivel", "Codigo", "Descripcion", "Und", "Cant. Sistema", "Observaciones"]
        ws.append(headers)
        
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    @staticmethod
    def generar_plantilla_transito() -> BytesIO:
        """Genera un buffer con la plantilla Excel de Tránsito."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transito_Detallado"
        # Cabeceras: Codigo / SKU | Documento | Cantidad
        headers = ["Codigo / SKU", "Documento", "Cantidad"]
        ws.append(headers)
        
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    @staticmethod
    async def importar_legacy_excel(
        file_content: bytes, 
        ronda: int,
        db: AsyncSession
    ) -> Dict[str, Any]:
        """
        Importa resultados de conteo del año pasado (Legacy).
        Cruce por B. Siigo (col 0) y Codigo (col 6).
        Valor en Cant. (col 10).
        """
        from io import BytesIO
        import openpyxl
        from sqlalchemy import and_, select
        
        wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        sheet = wb.active
        
        actualizados = 0
        errores = []
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            
            try:
                # Mapeo según estructura usuario:
                # col 0: B. Siigo
                # col 6: Codigo
                # col 10: Cant.
                b_siigo_val = int(float(row[0])) if row[0] is not None else None
                codigo_val = str(row[6]) if row[6] is not None else ""
                cantidad_val = float(row[10]) if row[10] is not None else 0.0
                
                # Buscar el registro exacto
                stmt = select(ConteoInventario).where(and_(
                    ConteoInventario.b_siigo == b_siigo_val,
                    ConteoInventario.codigo == codigo_val
                ))
                result = await db.execute(stmt)
                item = result.scalar_one_or_none()
                
                if item:
                    # Actualizar ronda específica
                    if ronda == 1:
                        item.cant_c1 = cantidad_val
                    elif ronda == 2:
                        item.cant_c2 = cantidad_val
                    elif ronda == 3:
                        item.cant_c3 = cantidad_val
                    
                    # Forzar recalculo de estado
                    teorico = (item.cantidad_sistema or 0.0) + (item.invporlegalizar or 0.0)
                    item.diferencia = cantidad_val - teorico
                    
                    if ronda == 1:
                        item.estado = "CONCILIADO" if cantidad_val == teorico else "RECONTEO"
                    
                    actualizados += 1
                else:
                    errores.append(f"Fila {row_idx}: No hallado {codigo_val} en Bodega {b_siigo_val}")
                    
            except Exception as e:
                errores.append(f"Fila {row_idx}: {str(e)}")
        
        if actualizados > 0:
            await db.commit()
            
        return {
            "exito": len(errores) == 0,
            "actualizados": actualizados,
            "errores": errores
        }

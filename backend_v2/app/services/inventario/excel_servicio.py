import openpyxl
from io import BytesIO
from typing import Dict, Any
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text, and_
from sqlmodel import select
from ...models.inventario.conteo import (
    ConteoInventario, 
    TransitoInventario, 
    ConteoHistorico
)

class ServicioExcelInventario:
    @staticmethod
    async def crear_snapshot(db: AsyncSession):
        """Copia todos los registros actuales a la tabla de historial antes de limpiar."""
        stmt = select(ConteoInventario)
        result = await db.execute(stmt)
        items = result.scalars().all()
        
        for item in items:
            hist = ConteoHistorico(
                original_id=item.id,
                b_siigo=item.b_siigo,
                bodega=item.bodega,
                bloque=item.bloque,
                estante=item.estante,
                nivel=item.nivel,
                codigo=item.codigo,
                descripcion=item.descripcion,
                unidad=item.unidad,
                cantidad_sistema=item.cantidad_sistema,
                cant_c1=item.cant_c1,
                cant_c2=item.cant_c2,
                cant_c3=item.cant_c3,
                cant_c4=item.cant_c4,
                conteo=item.conteo,
                estado=item.estado,
                invporlegalizar=item.invporlegalizar,
                cantidad_final=item.cantidad_final,
                diferencia_total=item.diferencia_total
            )
            db.add(hist)
        
        await db.flush()

    @staticmethod
    async def importar_conteo_excel(
        file_content: bytes, 
        nombre_conteo: str,
        db: AsyncSession,
        limpiar_previo: bool = False
    ) -> Dict[str, Any]:
        """Procesa un archivo Excel y carga los registros en inventario_conteo."""
        wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        sheet = wb.active
        
        registros_creados = 0
        errores = []

        if limpiar_previo:
            await ServicioExcelInventario.crear_snapshot(db)
            await db.execute(text("TRUNCATE TABLE conteoinventario RESTART IDENTITY CASCADE"))

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): 
                continue
            
            try:
                bodega_val = str(row[1]).strip() if row[1] is not None else ""
                bloque_val = str(row[2]).strip() if row[2] is not None else ""
                estante_val = str(row[3]).strip() if row[3] is not None else ""
                nivel_val = str(row[4]).strip() if row[4] is not None else ""
                codigo_val = str(row[5]).strip() if row[5] is not None else ""
                descripcion_val = str(row[6]).strip() if row[6] is not None else ""
                cant_sist_val = float(row[8]) if row[8] is not None else 0.0
                legalizar_val = float(row[9]) if len(row) > 9 and row[9] is not None else 0.0

                if not limpiar_previo:
                    stmt = select(ConteoInventario).where(and_(
                        ConteoInventario.codigo == codigo_val,
                        ConteoInventario.bodega == bodega_val,
                        ConteoInventario.bloque == bloque_val,
                        ConteoInventario.estante == estante_val,
                        ConteoInventario.nivel == nivel_val
                    ))
                    item_existente = (await db.execute(stmt)).scalar_one_or_none()
                    
                    if item_existente:
                        item_existente.cantidad_sistema = cant_sist_val
                        item_existente.invporlegalizar = legalizar_val
                        cant_f = cant_sist_val + legalizar_val
                        item_existente.cantidad_final = cant_f
                        if item_existente.cant_c1 is None and item_existente.cant_c2 is None:
                            item_existente.diferencia_total = -cant_f
                            
                        item_existente.conteo = nombre_conteo
                        item_existente.descripcion = descripcion_val
                        item_existente.unidad = str(row[7]) if row[7] is not None else ""
                        if item_existente.estado == "CONCILIADO":
                            item_existente.estado = "PENDIENTE"
                        registros_creados += 1
                        continue

                conteo = ConteoInventario(
                    b_siigo=int(row[0]) if row[0] is not None and str(row[0]).isdigit() else None,
                    bodega=bodega_val, bloque=bloque_val, estante=estante_val, nivel=nivel_val,
                    codigo=codigo_val, descripcion=descripcion_val, unidad=str(row[7]) if row[7] is not None else "",
                    cantidad_sistema=cant_sist_val, invporlegalizar=legalizar_val,
                    cantidad_final=cant_sist_val + legalizar_val, diferencia_total=-(cant_sist_val + legalizar_val),
                    conteo=nombre_conteo
                )
                db.add(conteo)
                registros_creados += 1
            except Exception as e:
                errores.append(f"Fila {row_idx}: {str(e)}")
        
        if registros_creados > 0:
            await db.commit()
            
        return {"exito": len(errores) == 0, "creados": registros_creados, "errores": errores}

    @staticmethod
    def _parse_float(val: Any) -> float:
        if val is None: 
            return 0.0
        if isinstance(val, (int, float)): 
            return float(val)
        try:
            return float(str(val).replace(',', '.').replace(' ', '').strip())
        except (ValueError, TypeError): 
            return 0.0

    @staticmethod
    async def importar_transito_excel(file_content: bytes, db: AsyncSession) -> Dict[str, Any]:
        wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        sheet = wb.active
        registros_creados = 0
        errores = []
        await db.execute(text("TRUNCATE TABLE transitoinventario"))

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): 
                continue
            try:
                sku_val = str(row[0]).strip() if row[0] is not None else ""
                doc_val = str(row[1]).strip() if len(row) > 1 and row[1] is not None else "TRANSITO_S_D"
                cant_val = ServicioExcelInventario._parse_float(row[2]) if len(row) > 2 else 0.0

                transito = TransitoInventario(sku=sku_val, documento=doc_val, cantidad=cant_val)
                db.add(transito)
                registros_creados += 1
            except Exception as e:
                errores.append(f"Fila {row_idx}: {str(e)}")
        
        if registros_creados > 0:
            await db.commit()
            await ServicioExcelInventario.sincronizar_transito_con_maestra(db)
            
        return {"exito": len(errores) == 0, "creados": registros_creados, "errores": errores}

    @staticmethod
    async def sincronizar_transito_con_maestra(db: AsyncSession):
        res_sum = await db.execute(text("SELECT sku, SUM(cantidad) FROM transitoinventario GROUP BY sku"))
        transito_dict = {row[0]: float(row[1]) for row in res_sum.fetchall()}
        await db.execute(text("UPDATE conteoinventario SET invporlegalizar = 0"))
        
        for sku, total in transito_dict.items():
            await db.execute(text("""
                UPDATE conteoinventario SET invporlegalizar = :total, 
                cantidad_final = cantidad_sistema + :total, diferencia_total = -(cantidad_sistema + :total)
                WHERE codigo = :sku
            """), {"total": total, "sku": sku})
        await db.commit()

    @staticmethod
    def generar_plantilla_maestra() -> BytesIO:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Maestra_Inventario"
        ws.append(["B. Siigo", "Bodega", "Bloque", "Estante", "Nivel", "Codigo", "Descripcion", "Und", "Cant. Sistema", "Observaciones"])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    @staticmethod
    def generar_plantilla_transito() -> BytesIO:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transito_Detallado"
        ws.append(["Codigo / SKU", "Documento", "Cantidad"])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    @staticmethod
    async def importar_legacy_excel(file_content: bytes, ronda: int, db: AsyncSession) -> Dict[str, Any]:
        wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        sheet = wb.active
        actualizados = 0
        errores = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): 
                continue
            try:
                b_siigo_val = int(float(row[0])) if row[0] is not None else None
                codigo_val = str(row[6]) if row[6] is not None else ""
                cantidad_val = float(row[10]) if row[10] is not None else 0.0
                
                stmt = select(ConteoInventario).where(and_(
                    ConteoInventario.b_siigo == b_siigo_val, ConteoInventario.codigo == codigo_val
                ))
                item = (await db.execute(stmt)).scalar_one_or_none()
                if item:
                    setattr(item, f"cant_c{ronda}", cantidad_val)
                    teorico = (item.cantidad_sistema or 0.0) + (item.invporlegalizar or 0.0)
                    item.diferencia = cantidad_val - teorico
                    if ronda == 1: 
                        item.estado = "CONCILIADO" if cantidad_val == teorico else "RECONTEO"
                    actualizados += 1
                else: 
                    errores.append(f"Fila {row_idx}: No hallado {codigo_val}")
            except Exception as e: 
                errores.append(f"Fila {row_idx}: {str(e)}")
        if actualizados > 0: 
            await db.commit()
        return {"exito": len(errores) == 0, "actualizados": actualizados, "errores": errores}

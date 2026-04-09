from typing import List, Dict, Optional
from sqlalchemy.orm import Session
from sqlalchemy import text
from datetime import date
import uuid
import json
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from .viaticos_query_service import ViaticosQueryService

class ViaticosService:
    """Lógica de comandos (escritura) para viáticos en el ERP (Solid)"""

    @staticmethod
    def enviar_reporte(db_erp: Session, reporte_data: Dict) -> str:
        """Guarda o actualiza un reporte de viáticos en la tabla de tránsito del ERP"""

        # Generar o Recuperar Consecutivo Amigable (Formato WEB-LXXXX)
        try:
            reporte_id_val = reporte_data.get("reporte_id")
            es_actualizacion = bool(
                reporte_id_val
                and isinstance(reporte_id_val, str)
                and "WEB-L" in reporte_id_val
                and reporte_id_val.strip().lower() != "null"
            )

            if es_actualizacion:
                consecutivo = str(reporte_id_val).strip()
            else:
                # Generación robusta: buscar el último número usado
                sql_max = text("""
                    SELECT reporte_id FROM legalizaciones_transito 
                    WHERE reporte_id LIKE 'WEB-L%' 
                    ORDER BY length(reporte_id) DESC, reporte_id DESC 
                    LIMIT 1
                """)
                max_val = db_erp.execute(sql_max).scalar()

                if max_val:
                    try:
                        num_str = "".join(filter(str.isdigit, max_val))
                        nuevo_num = int(num_str) + 1 if num_str else 1
                    except Exception:
                        sql_count = text("SELECT COUNT(*) FROM legalizaciones_transito")
                        nuevo_num = (db_erp.execute(sql_count).scalar() or 0) + 1
                else:
                    nuevo_num = 1

                consecutivo = f"WEB-L{nuevo_num:04d}"
        except Exception as e:
            print(f"DEBUG_ID_GEN_ERROR: {e}")
            consecutivo = f"WEB-L{uuid.uuid4().hex[:6].upper()}"

        # Si es actualización, VALIDAR ESTADO y luego LIMPIAR líneas previas
        if es_actualizacion:
            try:
                # Blindaje: No permitir modificar si ya está PROCESADO
                sql_check = text(
                    "SELECT estado FROM legalizaciones_transito WHERE reporte_id = :rid"
                )
                estado_actual = db_erp.execute(sql_check, {"rid": reporte_id_val}).scalar()

                if estado_actual and estado_actual not in ["BORRADOR", "INICIAL"]:
                    print(f"SECURITY_ALERT: Intento de modificar reporte bloqueado {reporte_id_val} en estado {estado_actual}")
                    raise Exception(f"Este reporte se encuentra bloqueado (Estado: {estado_actual}). Ya no es posible realizar cambios.")

                db_erp.execute(text("DELETE FROM transito_viaticos WHERE reporte_id = :rid"), {"rid": reporte_id_val})
                db_erp.execute(text("DELETE FROM legalizaciones_transito WHERE reporte_id = :rid"), {"rid": reporte_id_val})
            except Exception as e:
                print(f"ERROR ERP Limpieza: {e}")
                db_erp.rollback()
                raise e

        reporte_id = consecutivo
        obs_gral = reporte_data.get("observaciones_gral", "")

        # Evitar duplicación de etiquetas
        import re
        obs_gral = re.sub(r"\[WEB-L\d+\]\s*", "", obs_gral).strip()

        # Calcular Totales y Anexos
        total_acumulado = sum(
            float(g["valorConFactura"] or 0) + float(g["valorSinFactura"] or 0)
            for g in reporte_data["gastos"]
        )
        tiene_anexos = 1 if any(len(g.get("adjuntos", [])) > 0 for g in reporte_data["gastos"]) else 0
        cc_empleado = reporte_data.get("centrocosto") or "POR-DEFINIR"

        # Inserción de cabecera y detalles en una sola transacción
        try:
            # Limpiar cédula
            clean_uid = "".join(filter(str.isdigit, str(reporte_data.get("usuario_id", "0"))))
            usuario_int = int(clean_uid) if clean_uid else 0

            sql_header = text("""
                INSERT INTO legalizaciones_transito (
                    codigolegalizacion, fecha, hora, fechaaplicacion, empleado, nombreempleado, 
                    area, valortotal, estado, usuario, observaciones, 
                    anexo, centrocosto, cargo, ciudad, reporte_id
                ) VALUES (
                    :codigolegalizacion, CURRENT_DATE, CURRENT_TIME, CURRENT_DATE, :empleado, :nombreempleado,
                    :area, :valortotal, :estado, :usuario, :observaciones,
                    :anexo, :centrocosto, :cargo, :ciudad, :reporte_id
                ) RETURNING codigo
            """)

            result_header = db_erp.execute(
                sql_header,
                {
                    "codigolegalizacion": str(reporte_id),
                    "empleado": str(reporte_data["empleado_cedula"]),
                    "nombreempleado": str(reporte_data["empleado_nombre"]),
                    "area": str(reporte_data["area"]),
                    "valortotal": float(total_acumulado),
                    "estado": str(reporte_data.get("estado", "INICIAL")),
                    "usuario": usuario_int,
                    "observaciones": obs_gral.strip(),
                    "anexo": int(tiene_anexos),
                    "centrocosto": str(cc_empleado if cc_empleado and cc_empleado != "---" else "POR-DEFINIR"),
                    "cargo": str(reporte_data["cargo"]),
                    "ciudad": str(reporte_data["ciudad"]),
                    "reporte_id": str(reporte_id),
                },
            )

            cabecera_id_numerico = result_header.scalar()

            sql_insert = text("""
                INSERT INTO transito_viaticos (
                    legalizacion, fecha, fecharealgasto, categoria, ot, 
                    centrocosto, subcentrocosto, valorconfactura, valorsinfactura, 
                    observaciones, reporte_id, estado, fecha_registro, 
                    empleado_cedula, empleado_nombre, area, cargo, ciudad, 
                    observaciones_gral, usuario_id, adjuntos
                ) VALUES (
                    :legalizacion, :fecha, :fecharealgasto, :categoria, :ot, 
                    :centrocosto, :subcentrocosto, :valorconfactura, :valorsinfactura, 
                    :observaciones, :reporte_id, :estado, CURRENT_TIMESTAMP, 
                    :empleado_cedula, :empleado_nombre, :area, :cargo, :ciudad, 
                    :observaciones_gral, :usuario_id, CAST(:adjuntos AS jsonb)
                )
            """)

            for gasto in reporte_data["gastos"]:
                db_erp.execute(
                    sql_insert,
                    {
                        "legalizacion": cabecera_id_numerico,
                        "fecha": date.today(),
                        "fecharealgasto": gasto["fecha"],
                        "categoria": str(gasto["categoria"]),
                        "ot": str(gasto["ot"]),
                        "centrocosto": str(gasto["cc"]),
                        "subcentrocosto": str(gasto["scc"]),
                        "valorconfactura": float(gasto["valorConFactura"] or 0),
                        "valorsinfactura": float(gasto["valorSinFactura"] or 0),
                        "observaciones": str(gasto.get("observaciones", "")),
                        "reporte_id": str(reporte_id),
                        "estado": str(reporte_data.get("estado", "INICIAL")),
                        "empleado_cedula": str(reporte_data["empleado_cedula"]),
                        "empleado_nombre": str(reporte_data["empleado_nombre"]),
                        "area": str(reporte_data["area"]),
                        "cargo": str(reporte_data["cargo"]),
                        "ciudad": str(reporte_data["ciudad"]),
                        "observaciones_gral": str(obs_gral),
                        "usuario_id": usuario_int,
                        "adjuntos": json.dumps(gasto.get("adjuntos", [])),
                    },
                )

            db_erp.commit()
            print(f"REPORT_SUCCESS | ID: {reporte_id} | Total: {total_acumulado}")
            return str(reporte_id)

        except Exception as e:
            db_erp.rollback()
            print(f"CRITICAL_ERP_ERROR: {str(e)}")
            raise e

    @staticmethod
    def eliminar_reporte(db_erp: Session, reporte_id: str):
        """Elimina físicamente un reporte y sus líneas de las tablas de tránsito"""
        try:
            # Validar estado actual antes de eliminar
            sql_estado = text("SELECT estado FROM legalizaciones_transito WHERE reporte_id = :rid")
            estado_actual = db_erp.execute(sql_estado, {"rid": reporte_id}).scalar()

            if estado_actual and estado_actual.upper().strip() == "PROCESADO":
                raise ValueError(f"El reporte {reporte_id} ya fue PROCESADO y no puede ser eliminado.")

            db_erp.execute(text("DELETE FROM transito_viaticos WHERE reporte_id = :rid"), {"rid": reporte_id})
            db_erp.execute(text("DELETE FROM legalizaciones_transito WHERE reporte_id = :rid"), {"rid": reporte_id})
            db_erp.commit()
            print(f"REPORT_DELETE_SUCCESS | ID: {reporte_id}")
        except Exception as e:
            db_erp.rollback()
            raise e

    @staticmethod
    def exportar_estado_cuenta_xlsx(
        db_erp: Session,
        cedula: str,
        desde: Optional[date] = None,
        hasta: Optional[date] = None,
    ) -> io.BytesIO:
        """Genera un archivo XLSX con el estado de cuenta"""
        try:
            # Reutiliza el servicio de consultas
            data = ViaticosQueryService.obtener_estado_cuenta(db_erp, cedula, desde, hasta)

            wb = Workbook()
            ws = wb.active
            ws.title = "Estado de Cuenta"

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

            headers = ["Fecha Aplicación", "Radicado", "Tipo", "Observaciones", "Consignaciones (Contab.)", "Legalizaciones (Contab.)", "Consignaciones (Firmas)", "Legalizaciones (Firmas)", "Consignaciones (Pend.)", "Legalizaciones (Pend.)", "Saldo"]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            for row_idx, item in enumerate(data, 2):
                ws.cell(row=row_idx, column=1, value=item["fechaaplicacion"]).border = border
                ws.cell(row=row_idx, column=2, value=item["radicado"]).border = border
                ws.cell(row=row_idx, column=3, value=item["tipo"]).border = border
                ws.cell(row=row_idx, column=4, value=item["observaciones"]).border = border

                cols_money = ["consignacion_contabilizado", "legalizacion_contabilizado", "consignacion_firmadas", "legalizacion_firmadas", "consignacion_pendientes", "legalizacion_pendientes", "saldo"]
                for i, key in enumerate(cols_money, 5):
                    cell = ws.cell(row=row_idx, column=i, value=float(item[key] or 0))
                    cell.number_format = '"$"#,##0'
                    cell.border = border

            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 20

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output
        except Exception as e:
            print(f"ERROR ERP Export XLSX: {e}")
            raise e

import io
import base64
import asyncio
import gzip
import zipfile
from types import SimpleNamespace
from unittest.mock import AsyncMock, MagicMock, patch

import openpyxl
import pytest
from fastapi import HTTPException

from app.services.novedades_nomina.nomina_service import NominaService
from app.services.novedades_nomina.validacion_excel_hdi import (
    ArchivoHdiInvalido,
    MAX_EXCEL_COLS,
    MAX_EXCEL_ROWS,
    MAX_EXCEL_SHEETS,
    validar_excel_hdi,
)


XLS_BIFF8_GZIP_BASE64 = (
    "H4sIAAAAAAAAA+2X30uTURjHP+fs3dwPne+2khSqMYggIzenpheamy6c6DY2I4IglciF"
    "iTG96S76cVfQH9BNl11VNwXRhf0BUXjZlfYXFHhZLs6zzaYQuARB2hfOs/N+z4/n+77"
    "POc85+/I5sPHiTdcmezCCg+2yB1cdpwBde7Arz6Zsl8vlGl1u4kjhV7lc3q4WC3BW49"
    "ks/08xe9oB1OJv9nwL4AY8gBfwAa1AG+AH2ispgAAQBELAMeA40NFcQ0em5Pcm/gZh4"
    "VS1XKL2OcastTVn7UwpUCROnFUm6W/YfxAt/nX17Nm/7gquskyJReZZZpnFf/Zfw37Gm"
    "O+UcTbs6q9o1H89LvYO9B7Et8dtg8vJfOlTi2dN4WqBTTSvrY9yV/gGXOeuySeFYjy+"
    "Olk4iLf9Iyka5pTRsKZgWGkUz9EE6BRlQbEhsa+k3wexw5LPYLa0Nj4oNcWWGpV+T8V"
    "GxPoxM76TMV+FidHFdxnxrLoVnCpBidvMcYdTVj+DnEBF8GYmCj34MhOFMBGKFImQvB"
    "VhiSUiuPIRVlghwvo9074hiRW29gS3yR8uvxezEvrH+Ok2GQ2HmmCcNE+wUCh+ygFag"
    "Xdnq9i7tkqbLKJWNDdpl3pAlpKNm58vf6xPz+cu3RD+nPDdYh8Ic79u/ofCP6pcz7Vf"
    "uMpV3eQZSwekbhtlY6TIM1MlVYWcIU2OLJY+IyNt/Kg046TIYBovk2aMBGOkyZLB0j2"
    "SyW1CqAxZpkmSJ0WBMNcIkyBHiimmMNNkKWDp05J4bXyoHHnSTJOQrhmukGAKS7dJdr"
    "TNC8aIYmmffEvb2ByW7kSJPDdqiD76iDFAlErffpRI6kBNissUYfJkSZImQ0GUh8mJU"
    "PMmCZFm6ZMoEeZFxRkixgWiDNLPeaJEGdFB3so/odG6+L9XHhxUg77gDrHwJx6mwfwU"
    "3Wcp7qIFauf8UfJJrOoSOXz8Bq07/6sADgAA"
)


def _xlsx_bytes(*, rows=1, cols=5, sheets=1):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "HDI_1"
    for sheet_index in range(1, sheets):
        workbook.create_sheet(f"HDI_{sheet_index + 1}")
    for sheet in workbook.worksheets:
        for row in range(1, rows + 1):
            for col in range(1, cols + 1):
                sheet.cell(row=row, column=col, value="dato")
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _upload_xlsx():
    upload = AsyncMock()
    upload.filename = "hdi.xlsx"
    upload.content_type = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    upload.read = AsyncMock(return_value=_xlsx_bytes())
    return upload


def test_validar_excel_hdi_rechaza_pdf_disfrazado_y_zip_generico():
    with pytest.raises(ArchivoHdiInvalido):
        validar_excel_hdi(b"%PDF-1.4", "nomina.xlsx")

    output = io.BytesIO()
    with zipfile.ZipFile(output, "w") as archive:
        archive.writestr("archivo.txt", "no es Excel")
    with pytest.raises(ArchivoHdiInvalido):
        validar_excel_hdi(output.getvalue(), "nomina.xlsx")


def test_validar_excel_hdi_rechaza_extension_contenedor_incompatibles():
    xlsx = _xlsx_bytes()
    with pytest.raises(ArchivoHdiInvalido):
        validar_excel_hdi(xlsx, "nomina.xls")

    with pytest.raises(ArchivoHdiInvalido):
        validar_excel_hdi(bytes.fromhex("D0CF11E0A1B11AE1") + b"invalido", "nomina.xlsx")


def test_validar_y_extraer_archivo_xls_biff8_real():
    from app.services.novedades_nomina.hdi_extractor import extraer_hdi

    contenido = gzip.decompress(base64.b64decode(XLS_BIFF8_GZIP_BASE64))
    assert validar_excel_hdi(contenido, "hdi.xls", "application/vnd.ms-excel") == "xls"

    rows, summary, _warnings = extraer_hdi([contenido])
    assert summary["archivos_procesados"] == 1
    assert rows[0]["cedula"] == "94416010"


def test_api_hdi_requiere_permiso_y_rechaza_multiples_archivos():
    from fastapi.testclient import TestClient

    from app.api.novedades_nomina.dependencies import requiere_permiso_nomina_novedades
    from app.core.rate_limiter import limiter
    from app.main import app

    limiter.enabled = False
    client = TestClient(app)
    try:
        unauthorized = client.get(
            "/api/v2/novedades-nomina/hdi/datos",
            params={"mes": 7, "anio": 2026},
        )
        assert unauthorized.status_code in (401, 403)

        app.dependency_overrides[requiere_permiso_nomina_novedades] = (
            lambda: {"sub": "test_user"}
        )
        response = client.post(
            "/api/v2/novedades-nomina/hdi/preview",
            data={"mes": 7, "anio": 2026},
            files=[
                ("files", ("uno.xlsx", b"PK", "application/octet-stream")),
                ("files", ("dos.xlsx", b"PK", "application/octet-stream")),
            ],
        )
        assert response.status_code == 400
        assert "exactamente un archivo" in response.json()["detail"]
    finally:
        limiter.enabled = True
        app.dependency_overrides.clear()


@pytest.mark.parametrize(
    ("rows", "cols", "sheets", "mensaje"),
    [
        (MAX_EXCEL_ROWS + 1, 1, 1, "filas"),
        (1, MAX_EXCEL_COLS + 1, 1, "columnas"),
        (1, 1, MAX_EXCEL_SHEETS + 1, "hojas"),
    ],
)
def test_validar_excel_hdi_rechaza_limites_sin_truncar(rows, cols, sheets, mensaje):
    with pytest.raises(ArchivoHdiInvalido, match=mensaje):
        validar_excel_hdi(_xlsx_bytes(rows=rows, cols=cols, sheets=sheets), "nomina.xlsx")


@pytest.mark.asyncio
async def test_procesar_flujo_no_guarda_ni_elimina_si_extraccion_falla(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    session = AsyncMock()
    upload = _upload_xlsx()

    def extractor(_archivos):
        raise ValueError("libro corrupto")

    with pytest.raises(HTTPException) as exc_info:
        await NominaService.procesar_flujo(
            session=session,
            db_erp=AsyncMock(),
            files=[upload],
            categoria="OTROS",
            subcategoria="SEGUROS HDI",
            extractor_fn=extractor,
            extension="xlsx",
            mes=7,
            anio=2026,
        )

    assert exc_info.value.status_code == 422
    assert "libro corrupto" not in exc_info.value.detail
    session.execute.assert_not_awaited()
    session.commit.assert_not_awaited()
    assert not (tmp_path / "uploads").exists()


@pytest.mark.asyncio
async def test_procesar_flujo_preserva_periodo_si_extraccion_esta_vacia(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    session = AsyncMock()
    upload = _upload_xlsx()

    with pytest.raises(HTTPException, match="preservar los datos existentes"):
        await NominaService.procesar_flujo(
            session=session,
            db_erp=AsyncMock(),
            files=[upload],
            categoria="OTROS",
            subcategoria="SEGUROS HDI",
            extractor_fn=lambda _archivos: ([], {}, ["sin filas"]),
            extension="xlsx",
            mes=7,
            anio=2026,
        )

    session.execute.assert_not_awaited()
    assert not (tmp_path / "uploads").exists()


@pytest.mark.asyncio
@pytest.mark.parametrize("error", [RuntimeError("fallo DB"), asyncio.CancelledError()])
async def test_procesar_flujo_limpia_archivo_si_falla_despues_de_escribir(
    tmp_path, monkeypatch, error
):
    monkeypatch.chdir(tmp_path)
    session = AsyncMock()
    session.execute.side_effect = [MagicMock(), error]
    upload = _upload_xlsx()
    rows = [{"cedula": "94416010", "valor": 100.0}]

    with (
        patch(
            "app.services.novedades_nomina.nomina_service.ExcepcionService.obtener_excepciones_activas",
            new=AsyncMock(return_value=[]),
        ),
        patch.object(NominaService, "get_mapa_erp", new=AsyncMock(return_value={})),
        pytest.raises(type(error)),
    ):
        await NominaService.procesar_flujo(
            session=session,
            db_erp=AsyncMock(),
            files=[upload],
            categoria="OTROS",
            subcategoria="SEGUROS HDI",
            extractor_fn=lambda _archivos: (rows, {}, []),
            extension="xlsx",
            mes=7,
            anio=2026,
        )

    session.rollback.assert_awaited_once()
    storage = tmp_path / "uploads" / "nomina"
    assert storage.exists()
    assert [path for path in storage.iterdir() if path.name != ".quota.lock"] == []


@pytest.mark.asyncio
async def test_cancelacion_durante_commit_no_elimina_archivo_confirmado(
    tmp_path, monkeypatch
):
    monkeypatch.chdir(tmp_path)
    session = AsyncMock()
    session.execute.side_effect = [MagicMock(), MagicMock()]
    commit_iniciado = asyncio.Event()
    permitir_commit = asyncio.Event()

    async def commit_demorado():
        commit_iniciado.set()
        await permitir_commit.wait()

    session.commit.side_effect = commit_demorado
    upload = _upload_xlsx()
    registro = SimpleNamespace(
        cedula="94416010",
        nombre_asociado="PRECIADO JOSE",
        valor=100.0,
        valor_rdc=24.0,
        valor_colaborador=76.0,
        empresa="REFRIDCOL",
        concepto="SEGURO DE VIDA",
        ciudad=None,
        observaciones=None,
        horas=0,
        dias=0,
        estado_validacion="OK",
    )

    with (
        patch(
            "app.services.novedades_nomina.nomina_service.ExcepcionService.obtener_excepciones_activas",
            new=AsyncMock(return_value=[]),
        ),
        patch.object(NominaService, "get_mapa_erp", new=AsyncMock(return_value={})),
        patch.object(
            NominaService,
            "crear_archivo_procesado",
            new=AsyncMock(return_value=SimpleNamespace(id=7)),
        ),
        patch.object(
            NominaService,
            "persistir_registros_normalizados",
            new=AsyncMock(return_value=[registro]),
        ),
    ):
        tarea = asyncio.create_task(
            NominaService.procesar_flujo(
                session=session,
                db_erp=AsyncMock(),
                files=[upload],
                categoria="OTROS",
                subcategoria="SEGUROS HDI",
                extractor_fn=lambda _archivos: ([{"cedula": "94416010"}], {}, []),
                extension="xlsx",
                mes=7,
                anio=2026,
            )
        )
        await commit_iniciado.wait()
        tarea.cancel()
        permitir_commit.set()
        with pytest.raises(asyncio.CancelledError):
            await tarea

    storage = tmp_path / "uploads" / "nomina"
    assert len([
        path for path in storage.iterdir() if path.name != ".quota.lock"
    ]) == 1
